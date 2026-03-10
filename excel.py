import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# colores
filled = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
empty = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# cargar json
with open("test_cases_100.json") as f:
    tests = json.load(f)

wb = Workbook()
wb.remove(wb.active)

for name, board in tests.items():

    ws = wb.create_sheet(title=name)

    for r in range(len(board)):
        for c in range(len(board[r])):

            cell = ws.cell(row=r+1, column=c+1)
            value = board[r][c]

            cell.value = value

            if value == 1:
                cell.fill = filled
            else:
                cell.fill = empty

            ws.column_dimensions[chr(65+c)].width = 3
        ws.row_dimensions[r+1].height = 18

wb.save("tetris_test_cases.xlsx")

print("Archivo creado: tetris_test_cases.xlsx")